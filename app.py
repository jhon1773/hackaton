import os
from flask import Flask
from flask_sqlalchemy import SQLAlchemy

app = Flask(__name__)

# Configuración de la base de datos para Render
DATABASE_URL = os.environ.get('DATABASE_URL')
if DATABASE_URL:
    # Render puede entregar la URL con 'postgres://' en vez de 'postgresql://'
    DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://')
    app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
else:
    # Configuración local (ajusta según tu entorno)
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///local.db'

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

from flask import Flask, render_template, request, flash, redirect, url_for, jsonify, session
from config import Config
from models import db, PQRSD, Usuario, Area, Historial, Rol
from datetime import datetime, timedelta
from sqlalchemy import func, extract

app = Flask(__name__)
app.config.from_object(Config)
app.secret_key = 'supersecretkey'  # Cambia esto por una clave segura

# Inicializar SQLAlchemy
db.init_app(app)

@app.route('/logout')
def logout():
    session.clear()
    flash('Sesión cerrada correctamente', 'success')
    return redirect(url_for('login'))

with app.app_context():
    db.create_all()

    if not Area.query.first():
        areas_iniciales = [
            Area(nombre="Atención al Ciudadano", descripcion="Recepción inicial de PQRSD"),
            Area(nombre="Recursos Humanos", descripcion="PQRSD relacionadas con personal"),
            Area(nombre="Contabilidad", descripcion="PQRSD relacionadas con aspectos financieros"),
            Area(nombre="Jurídica", descripcion="PQRSD de carácter legal y denuncias"),
            Area(nombre="Tecnología", descripcion="PQRSD relacionadas con sistemas informáticos"),
        ]
        db.session.add_all(areas_iniciales)
        db.session.commit()

    if not Rol.query.first():
        roles_iniciales = [
            Rol(nombre="Administrador", descripcion="Encargado de supervisar y gestionar el proceso PQRSD"),
            Rol(nombre="Funcionario", descripcion="Responsable de responder PQRSD asignadas"),
            Rol(nombre="Usuario", descripcion="Persona externa que registra PQRSD"),
        ]
        db.session.add_all(roles_iniciales)
        db.session.commit()


@app.route("/")
def index():
    from models import PQRSD, Usuario
    
    # Obtener estadísticas reales
    total_pqrsd = PQRSD.query.count()
    pendientes = PQRSD.query.filter_by(estado="Pendiente").count()
    resueltas = PQRSD.query.filter_by(estado="Resuelta").count()
    usuarios_activos = Usuario.query.count()
    
    # Obtener áreas para el formulario
    areas = Area.query.all()
    tipos_peticionario = ["Persona natural", "Persona jurídica", "Autoridad pública", "Periodista", "Concejal", "Otro"]
    
    return render_template("index.html", 
                         total_pqrsd=total_pqrsd,
                         pendientes=pendientes,
                         resueltas=resueltas,
                         usuarios_activos=usuarios_activos,
                         areas=areas,
                         tipos_peticionario=tipos_peticionario)


@app.route("/dashboard")
def dashboard():
    # -------------------------------
    # Volumen por periodo (todos)
    # -------------------------------
    # MES (año/mes)
    q_mes = db.session.query(
        extract('year', PQRSD.fecha_creacion).label('anio'),
        extract('month', PQRSD.fecha_creacion).label('mes'),
        func.count(PQRSD.id).label('total')
    ).group_by('anio', 'mes').order_by('anio', 'mes').all()
    vol_mes = {f"{int(r.mes)}/{int(r.anio)}": r.total for r in q_mes}

    # TRIMESTRE (año - T1..T4)
    q_tri = db.session.query(
        extract('year', PQRSD.fecha_creacion).label('anio'),
        (func.floor((extract('month', PQRSD.fecha_creacion) - 1) / 3) + 1).label('tri'),
        func.count(PQRSD.id).label('total')
    ).group_by('anio', 'tri').order_by('anio', 'tri').all()
    vol_tri = {f"T{int(r.tri)}/{int(r.anio)}": r.total for r in q_tri}

    # SEMESTRE (año - S1..S2)
    q_sem = db.session.query(
        extract('year', PQRSD.fecha_creacion).label('anio'),
        (func.floor((extract('month', PQRSD.fecha_creacion) - 1) / 6) + 1).label('sem'),
        func.count(PQRSD.id).label('total')
    ).group_by('anio', 'sem').order_by('anio', 'sem').all()
    vol_sem = {f"S{int(r.sem)}/{int(r.anio)}": r.total for r in q_sem}

    # AÑO
    q_anio = db.session.query(
        extract('year', PQRSD.fecha_creacion).label('anio'),
        func.count(PQRSD.id).label('total')
    ).group_by('anio').order_by('anio').all()
    vol_anio = {f"{int(r.anio)}": r.total for r in q_anio}

    volumen = {"mes": vol_mes, "trimestre": vol_tri, "semestre": vol_sem, "año": vol_anio}

    # -------------------------------
    # Clasificación por tipo
    # -------------------------------
    tipo_data = db.session.query(PQRSD.tipo, func.count(PQRSD.id)) \
        .group_by(PQRSD.tipo).order_by(PQRSD.tipo).all()
    labels_tipo = [t[0] for t in tipo_data]
    data_tipo = [t[1] for t in tipo_data]

    # -------------------------------
    # Clasificación por dependencia
    # -------------------------------
    dep_data = db.session.query(Area.nombre, func.count(PQRSD.id)) \
        .join(Area, PQRSD.area_id == Area.id) \
        .group_by(Area.nombre).order_by(Area.nombre).all()
    labels_dep = [d[0] for d in dep_data]
    data_dep = [d[1] for d in dep_data]

    # -------------------------------
    # Cumplimiento de plazos (Resueltas)
    # -------------------------------
    en_plazo = db.session.query(PQRSD).filter(
    PQRSD.estado == "Resuelta",
    PQRSD.fecha_resolucion.isnot(None),
    PQRSD.fecha_resolucion <= PQRSD.fecha_limite
    ).count()

    extemporaneas = db.session.query(PQRSD).filter(
    PQRSD.estado == "Resuelta",
    PQRSD.fecha_resolucion.isnot(None),
    PQRSD.fecha_resolucion > PQRSD.fecha_limite
    ).count()

    cumplimiento = {"en_plazo": en_plazo, "extemporaneas": extemporaneas}

    # -------------------------------
    # Clasificación por tipo de peticionario (dinámico)
    # -------------------------------
    pet_data = db.session.query(PQRSD.tipo_peticionario, func.count(PQRSD.id)) \
        .group_by(PQRSD.tipo_peticionario).order_by(PQRSD.tipo_peticionario).all()
    labels_pet = [p[0] for p in pet_data]
    data_pet = [p[1] for p in pet_data]

    return render_template(
        "dashboard.html",
        volumen=volumen,
        labels_tipo=labels_tipo,
        data_tipo=data_tipo,
        labels_dep=labels_dep,
        data_dep=data_dep,
        cumplimiento=cumplimiento,
        labels_pet=labels_pet,
        data_pet=data_pet
    )


@app.route("/nueva_pqrsd", methods=["GET", "POST"])
def nueva_pqrsd():
    # Solo usuarios autenticados pueden registrar PQRSD
    if not session.get('user_id'):
        flash('Debes iniciar sesión para registrar una PQRSD.', 'warning')
        return redirect(url_for('login'))

    areas = Area.query.all()
    tipos_peticionario = ["Persona natural", "Persona jurídica", "Autoridad pública", "Periodista", "Concejal", "Otro"]

    if request.method == "POST":
        try:
            tipo = request.form.get("tipo")
            descripcion = request.form.get("descripcion")
            solicitante_nombre = request.form.get("solicitante_nombre")
            solicitante_identificacion = request.form.get("solicitante_identificacion")
            solicitante_contacto = request.form.get("solicitante_contacto")
            area_id = request.form.get("area_id")
            prioridad = request.form.get("prioridad")
            medio = request.form.get("medio")
            tipo_peticionario = request.form.get("tipo_peticionario")
            dias_limite = 15 if prioridad == "Baja" else (10 if prioridad == "Media" else 5)
            fecha_limite = datetime.now() + timedelta(days=dias_limite)

            nueva = PQRSD(
                tipo=tipo,
                descripcion=descripcion,
                solicitante_nombre=solicitante_nombre,
                solicitante_identificacion=solicitante_identificacion,
                solicitante_contacto=solicitante_contacto,
                area_id=area_id,
                prioridad=prioridad,
                fecha_limite=fecha_limite,
                estado="Pendiente",
                tipo_peticionario=tipo_peticionario,
                medio=medio
            )
            db.session.add(nueva)
            db.session.commit()

            flash("¡Su solicitud ha sido registrada exitosamente! Gracias por su sugerencia, reclamo, petición o denuncia.", "success")
            return redirect(url_for("nueva_pqrsd"))
        except Exception as e:
            db.session.rollback()
            flash(f"Error al crear PQRSD: {str(e)}", "danger")

    return render_template("nueva_pqrsd.html", areas=areas, tipos_peticionario=tipos_peticionario)


@app.route("/gestion")
def gestion():
    tipo = request.args.get("tipo", "")
    estado = request.args.get("estado", "")
    area_id = request.args.get("area_id", "")
    prioridad = request.args.get("prioridad", "")
    busqueda = request.args.get("busqueda", "")

    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 10, type=int)

    query = PQRSD.query
    if tipo:
        query = query.filter(PQRSD.tipo == tipo)
    if estado:
        estado_db = estado
        if estado == 'Resuelto':
            estado_db = 'Resuelta'
        query = query.filter(PQRSD.estado == estado_db)
    if area_id:
        query = query.filter(PQRSD.area_id == area_id)
    if prioridad:
        query = query.filter(PQRSD.prioridad == prioridad)

    # Filtro de búsqueda por cédula, nombre o radicado
    if busqueda:
        busqueda_like = f"%{busqueda}%"
        query = query.filter(
            (PQRSD.solicitante_identificacion.like(busqueda_like)) |
            (PQRSD.solicitante_nombre.like(busqueda_like)) |
            (PQRSD.id.like(busqueda_like))
        )

    items = query.order_by(PQRSD.fecha_creacion.desc()).all()
    return render_template("gestion.html", items=items)

@app.route("/responder_pqrs", methods=["GET", "POST"])
def responder_pqrs():
    pqrs_id = request.args.get("pqrs_id", "")
    if request.method == "POST":
        pqrs_id = request.form.get("pqrs_id")
        respuesta = request.form.get("respuesta")
        pqrs = PQRSD.query.filter_by(id=pqrs_id).first()
        if pqrs:
            pqrs.respuesta = respuesta
            pqrs.estado = "Resuelta"
            pqrs.fecha_resolucion = datetime.now()
            db.session.commit()
            flash("Respuesta enviada correctamente.", "success")
        else:
            flash("No se encontró la PQRS con ese ID.", "danger")
    return render_template("responder_pqrs.html", pqrs_id=pqrs_id)
    if request.method == "POST":
        pqrs_id = request.form.get("pqrs_id")
        respuesta = request.form.get("respuesta")
        pqrs = PQRSD.query.filter_by(id=pqrs_id).first()
        if pqrs:
            pqrs.respuesta = respuesta
            pqrs.estado = "Resuelta"
            pqrs.fecha_resolucion = datetime.now()
            db.session.commit()
            flash("Respuesta enviada correctamente.", "success")
        else:
            flash("No se encontró la PQRS con ese ID.", "danger")
    return render_template("responder_pqrs.html")

    tipo = request.args.get("tipo", "")
    estado = request.args.get("estado", "")
    area_id = request.args.get("area_id", "")
    prioridad = request.args.get("prioridad", "")
    busqueda = request.args.get("busqueda", "")

    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 10, type=int)

    query = PQRSD.query
    if tipo:
        query = query.filter(PQRSD.tipo == tipo)
    if estado:
        # Map frontend 'Resuelto' to backend 'Resuelta'
        estado_db = estado
        if estado == 'Resuelto':
            estado_db = 'Resuelta'
        query = query.filter(PQRSD.estado == estado_db)
    if area_id:
        query = query.filter(PQRSD.area_id == area_id)
    if prioridad:
        query = query.filter(PQRSD.prioridad == prioridad)
    if busqueda:
        search = f"%{busqueda}%"
        # Try exact match for radicado (ID) if busqueda is numeric
        if busqueda.isdigit():
            query = query.filter(
                (PQRSD.solicitante_identificacion.ilike(search)) |
                (PQRSD.solicitante_nombre.ilike(search)) |
                (PQRSD.id == int(busqueda))
            )
        else:
            query = query.filter(
                (PQRSD.solicitante_identificacion.ilike(search)) |
                (PQRSD.solicitante_nombre.ilike(search))
            )

    query = query.order_by(PQRSD.fecha_creacion.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    items = pagination.items

    areas = Area.query.all()

    return render_template(
        "gestion.html",
        items=items,
        pagination=pagination,
        areas=areas,
        filtros={"tipo": tipo, "estado": estado, "area_id": area_id, "prioridad": prioridad, "busqueda": busqueda}
    )


@app.route("/reportes")
def reportes():
    fecha_inicio = request.args.get("fecha_inicio", "")
    fecha_fin = request.args.get("fecha_fin", "")

    # Query base con filtros de fecha
    query = PQRSD.query
    if fecha_inicio:
        try:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            query = query.filter(PQRSD.fecha_creacion >= fecha_inicio_dt)
        except ValueError:
            pass
    if fecha_fin:
        try:
            fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d") + timedelta(days=1)
            query = query.filter(PQRSD.fecha_creacion < fecha_fin_dt)
        except ValueError:
            pass

    pqrsd_data = query.all()
    total = len(pqrsd_data)

    # Estadísticas por tipo
    por_tipo = db.session.query(PQRSD.tipo, func.count(PQRSD.id)) \
        .filter(PQRSD.id.in_([p.id for p in pqrsd_data])).group_by(PQRSD.tipo).all()
    por_tipo = dict(por_tipo)

    # Estadísticas por estado
    por_estado = db.session.query(PQRSD.estado, func.count(PQRSD.id)) \
        .filter(PQRSD.id.in_([p.id for p in pqrsd_data])).group_by(PQRSD.estado).all()
    por_estado = dict(por_estado)

    # Estadísticas por prioridad
    por_prioridad = db.session.query(PQRSD.prioridad, func.count(PQRSD.id)) \
        .filter(PQRSD.id.in_([p.id for p in pqrsd_data])).group_by(PQRSD.prioridad).all()
    por_prioridad = dict(por_prioridad)

    # Tiempo promedio de resolución
    pqrsd_resueltas = query.filter(PQRSD.estado == "Resuelta").all()
    tiempo_promedio = 0
    if pqrsd_resueltas:
        total_dias = sum([
            (p.fecha_resolucion - p.fecha_creacion).days
            for p in pqrsd_resueltas
            if p.fecha_resolucion is not None and p.fecha_creacion is not None
        ])
        tiempo_promedio = total_dias / len(pqrsd_resueltas)

    # Tendencias mensuales
    q_mes = db.session.query(
        extract('year', PQRSD.fecha_creacion).label('anio'),
        extract('month', PQRSD.fecha_creacion).label('mes'),
        func.count(PQRSD.id).label('total')
    )
    if fecha_inicio:
        q_mes = q_mes.filter(PQRSD.fecha_creacion >= fecha_inicio_dt)
    if fecha_fin:
        q_mes = q_mes.filter(PQRSD.fecha_creacion < fecha_fin_dt)
    q_mes = q_mes.group_by('anio', 'mes').order_by('anio', 'mes').all()
    tendencias_mensuales = [{
        'label': f"{int(r.mes)}/{int(r.anio)}",
        'total': r.total
    } for r in q_mes]

    # Estadísticas por área
    areas_stats = []
    areas = Area.query.all()
    for area in areas:
        total_area = query.filter(PQRSD.area_id == area.id).count()
        resueltas_area = query.filter(PQRSD.area_id == area.id, PQRSD.estado == "Resuelta").count()
        eficiencia = (resueltas_area / total_area * 100) if total_area > 0 else 0
        areas_stats.append({
            'nombre': area.nombre,
            'total': total_area,
            'resueltas': resueltas_area,
            'eficiencia': round(eficiencia, 1)
        })

    return render_template(
        "reportes.html",
        total=total,
        por_tipo=por_tipo,
        por_estado=por_estado,
        por_prioridad=por_prioridad,
        tiempo_promedio=tiempo_promedio,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        tendencias_mensuales=tendencias_mensuales,
        areas_stats=areas_stats
    )


@app.route("/api/pqrsd/<int:id>/cambiar_estado", methods=["POST"])
def cambiar_estado_pqrsd(id):
    try:
        pqrsd = PQRSD.query.get_or_404(id)
        nuevo_estado = request.json.get("estado")
        if nuevo_estado not in ["Pendiente", "En Proceso", "Resuelta", "Cerrada"]:
            return jsonify({"error": "Estado no válido"}), 400

        historial = Historial(
            pqrsd_id=id,
            usuario_id=1,  # TODO: reemplazar por el usuario autenticado
            accion=f"Cambio de estado: {pqrsd.estado} → {nuevo_estado}"
        )
        pqrsd.estado = nuevo_estado
        if nuevo_estado == "Resuelta":
            pqrsd.fecha_resolucion = datetime.now()

        db.session.add(historial)
        db.session.commit()

        return jsonify({"success": True, "nuevo_estado": nuevo_estado})

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500



# Ruta para ver detalle de una PQRSD
@app.route("/pqrsd/<int:pqrsd_id>")
def ver_pqrsd(pqrsd_id):
    pqrsd = PQRSD.query.get_or_404(pqrsd_id)
    return render_template("ver_pqrsd.html", pqrsd=pqrsd)

@app.route("/politica-privacidad")
def politica_privacidad():
    return render_template("politica_privacidad.html")

@app.route("/terminos-servicio")
def terminos_servicio():
    return render_template("terminos_servicio.html")

# Ruta para editar una PQRSD
@app.route("/pqrsd/<int:pqrsd_id>/editar", methods=["GET", "POST"])
def editar_pqrsd(pqrsd_id):
    pqrsd = PQRSD.query.get_or_404(pqrsd_id)
    if request.method == "POST":
        pqrsd.tipo = request.form.get("tipo", pqrsd.tipo)
        pqrsd.descripcion = request.form.get("descripcion", pqrsd.descripcion)
        pqrsd.prioridad = request.form.get("prioridad", pqrsd.prioridad)
        pqrsd.area_id = request.form.get("area_id", pqrsd.area_id)
        pqrsd.estado = request.form.get("estado", pqrsd.estado)
        db.session.commit()
        flash("PQRSD actualizada correctamente", "success")
        return redirect(url_for("ver_pqrsd", pqrsd_id=pqrsd.id))
    areas = Area.query.all()
    return render_template("editar_pqrsd.html", pqrsd=pqrsd, areas=areas)

@app.route('/login', methods=['GET', 'POST'])
def login():
    from models import Usuario
    if request.method == 'POST':
        usuario = request.form.get('usuario')
        password = request.form.get('password')
        user = Usuario.query.filter(Usuario.email == usuario).first()
        if user and user.password == password:
            session['user_id'] = user.id
            session['user_rol'] = user.rol.nombre
            flash('Inicio de sesión exitoso', 'success')
            # Redirigir según rol
            if user.rol.nombre == 'Administrador':
                return redirect(url_for('dashboard'))
            elif user.rol.nombre == 'Funcionario':
                return redirect(url_for('gestion'))
            else:
                return redirect(url_for('nueva_pqrsd'))
        else:
            flash('Usuario o contraseña incorrectos', 'danger')
    return render_template('login.html')
    # Solo Administrador puede ver el dashboard completo
    if 'user_rol' not in session or session['user_rol'] != 'Administrador':
        flash('Acceso restringido al dashboard', 'danger')
        return redirect(url_for('login'))
@app.route('/tabla_registro',)
def tabla_usuarios_registro():
        return render_template("tabla_usuarios_registro.html")

@app.route('/registro_usuario', methods=['GET', 'POST'])
def registro_usuario():
    from models import Usuario, Rol
    # Only allow 'Usuario' role in registration
    roles = Rol.query.filter_by(nombre='Usuario').all()
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        usuario_rol = Rol.query.filter_by(nombre='Usuario').first()
        rol_id = usuario_rol.id if usuario_rol else None
        # Validar que el correo no esté repetido
        if Usuario.query.filter_by(email=email).first():
            flash('El correo electrónico ya está registrado', 'danger')
        else:
            nuevo_usuario = Usuario(email=email, password=password, rol_id=rol_id)
            db.session.add(nuevo_usuario)
            db.session.commit()
            flash('Usuario registrado exitosamente', 'success')
            return redirect(url_for('login'))
    return render_template('registro_usuario.html', roles=roles)

@app.route('/exportar_excel')
def exportar_excel():
    from models import PQRSD, Area
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    pqrsd_list = PQRSD.query.all()
    data = []
    for p in pqrsd_list:
        area = Area.query.get(p.area_id)
        data.append({
            'ID': p.id,
            'Tipo': p.tipo,
            'Descripción': p.descripcion,
            'Solicitante': p.solicitante_nombre,
            'Identificación': p.solicitante_identificacion,
            'Contacto': p.solicitante_contacto,
            'Área': area.nombre if area else '',
            'Prioridad': p.prioridad,
            'Estado': p.estado,
            'Fecha Creación': p.fecha_creacion.strftime('%Y-%m-%d %H:%M'),
            'Fecha Límite': p.fecha_limite.strftime('%Y-%m-%d'),
            'Tipo Peticionario': p.tipo_peticionario,
            'Respuesta': p.respuesta if p.respuesta else ''
        })
    df = pd.DataFrame(data)
    from flask import send_file
    import io
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='PQRSD')
        workbook = writer.book
        worksheet = writer.sheets['PQRSD']
        # Estilos
        header_fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
    output.seek(0)
    return send_file(output, download_name='pqrsd.xlsx', as_attachment=True)

@app.route('/exportar_pdf')
def exportar_pdf():
    from models import PQRSD, Area
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    import io
    pqrsd_list = PQRSD.query.all()
    data = [['ID', 'Tipo', 'Descripción', 'Solicitante', 'Identificación', 'Contacto', 'Área', 'Prioridad', 'Estado', 'Fecha Creación', 'Fecha Límite', 'Tipo Peticionario', 'Respuesta']]
    for p in pqrsd_list:
        area = Area.query.get(p.area_id)
        data.append([
            p.id, p.tipo, p.descripcion, p.solicitante_nombre, p.solicitante_identificacion,
            p.solicitante_contacto, area.nombre if area else '', p.prioridad, p.estado,
            p.fecha_creacion.strftime('%Y-%m-%d %H:%M'), p.fecha_limite.strftime('%Y-%m-%d'), p.tipo_peticionario,
            p.respuesta if p.respuesta else ''
        ])
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(letter), leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20)
    style = getSampleStyleSheet()
    style_table = ParagraphStyle(name='TableCell', fontSize=8, leading=10, alignment=1, wordWrap='CJK')
    elements = [Paragraph('Reporte de PQRSD', style['Title']), Spacer(1, 12)]
    # Wrap text for all cells except header
    wrapped_data = [data[0]]
    for row in data[1:]:
        wrapped_row = [Paragraph(str(cell), style_table) for cell in row]
        wrapped_data.append(wrapped_row)
    table = Table(wrapped_data, repeatRows=1, colWidths=[30, 45, 100, 60, 50, 60, 50, 45, 45, 60, 60, 60])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1976D2')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.whitesmoke, colors.lightblue]),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elements.append(table)
    doc.build(elements)
    output.seek(0)
    from flask import send_file
    return send_file(output, download_name='pqrsd.pdf', as_attachment=True)

@app.route('/descargar_reporte_excel')
def descargar_reporte_excel():
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from models import PQRSD
    import pandas as pd
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    import io
    fecha_inicio = request.args.get("fecha_inicio", "")
    fecha_fin = request.args.get("fecha_fin", "")
    query = PQRSD.query
    if fecha_inicio:
        try:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            query = query.filter(PQRSD.fecha_creacion >= fecha_inicio_dt)
        except ValueError:
            pass
    if fecha_fin:
        try:
            fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d") + timedelta(days=1)
            query = query.filter(PQRSD.fecha_creacion < fecha_fin_dt)
        except ValueError:
            pass
    pqrsd_data = query.all()
    # Resumen por tipo, estado, prioridad
    resumen = []
    tipos = set([p.tipo for p in pqrsd_data])
    estados = set([p.estado for p in pqrsd_data])
    prioridades = set([p.prioridad for p in pqrsd_data])
    for tipo in tipos:
        count = sum(1 for p in pqrsd_data if p.tipo == tipo)
        resumen.append({'Indicador': 'Tipo', 'Valor': tipo, 'Cantidad': count})
    for estado in estados:
        count = sum(1 for p in pqrsd_data if p.estado == estado)
        resumen.append({'Indicador': 'Estado', 'Valor': estado, 'Cantidad': count})
    for prioridad in prioridades:
        count = sum(1 for p in pqrsd_data if p.prioridad == prioridad)
        resumen.append({'Indicador': 'Prioridad', 'Valor': prioridad, 'Cantidad': count})
    # Tendencias mensuales
    meses = {}
    for p in pqrsd_data:
        key = p.fecha_creacion.strftime('%m/%Y')
        meses[key] = meses.get(key, 0) + 1
    tendencias = pd.DataFrame({'Mes': list(meses.keys()), 'Total': list(meses.values())})
    df = pd.DataFrame(resumen)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Resumen')
        tendencias.to_excel(writer, index=False, sheet_name='Tendencias Mensuales')
        # Gráfica de tendencias
        plt.figure(figsize=(6,3))
        plt.bar(tendencias['Mes'], tendencias['Total'], color='#1976D2')
        plt.title('Tendencias Mensuales')
        plt.xlabel('Mes')
        plt.ylabel('Total')
        plt.tight_layout()
        img_bytes = io.BytesIO()
        plt.savefig(img_bytes, format='png')
        plt.close()
        img_bytes.seek(0)
        workbook = writer.book
        worksheet = workbook['Tendencias Mensuales']
        from openpyxl.drawing.image import Image as XLImage
        img = XLImage(img_bytes)
        worksheet.add_image(img, 'E2')
        # Estilos
        for ws in [workbook['Resumen'], worksheet]:
            header_fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = border
    output.seek(0)
    from flask import send_file
    return send_file(output, download_name='reporte_resumen.xlsx', as_attachment=True)

@app.route('/descargar_reporte_pdf')
def descargar_reporte_pdf():
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from models import PQRSD
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    import io
    fecha_inicio = request.args.get("fecha_inicio", "")
    fecha_fin = request.args.get("fecha_fin", "")
    query = PQRSD.query
    if fecha_inicio:
        try:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            query = query.filter(PQRSD.fecha_creacion >= fecha_inicio_dt)
        except ValueError:
            pass
    if fecha_fin:
        try:
            fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d") + timedelta(days=1)
            query = query.filter(PQRSD.fecha_creacion < fecha_fin_dt)
        except ValueError:
            pass
    pqrsd_data = query.all()
    # Resumen por tipo, estado, prioridad
    data = [['Indicador', 'Valor', 'Cantidad']]
    tipos = set([p.tipo for p in pqrsd_data])
    estados = set([p.estado for p in pqrsd_data])
    prioridades = set([p.prioridad for p in pqrsd_data])
    for tipo in tipos:
        count = sum(1 for p in pqrsd_data if p.tipo == tipo)
        data.append(['Tipo', tipo, count])
    for estado in estados:
        count = sum(1 for p in pqrsd_data if p.estado == estado)
        data.append(['Estado', estado, count])
    for prioridad in prioridades:
        count = sum(1 for p in pqrsd_data if p.prioridad == prioridad)
        data.append(['Prioridad', prioridad, count])
    # Tendencias mensuales
    meses = {}
    for p in pqrsd_data:
        key = p.fecha_creacion.strftime('%m/%Y')
        meses[key] = meses.get(key, 0) + 1
    tendencias_keys = list(meses.keys())
    tendencias_values = list(meses.values())
    plt.figure(figsize=(6,3))
    plt.bar(tendencias_keys, tendencias_values, color='#1976D2')
    plt.title('Tendencias Mensuales')
    plt.xlabel('Mes')
    plt.ylabel('Total')
    plt.tight_layout()
    img_bytes = io.BytesIO()
    plt.savefig(img_bytes, format='png')
    plt.close()
    img_bytes.seek(0)
    style = getSampleStyleSheet()
    style_table = ParagraphStyle(name='TableCell', fontSize=10, leading=12, alignment=1, wordWrap='CJK')
    elements = [Paragraph('Resumen de Reporte PQRSD', style['Title']), Spacer(1, 12)]
    elements.append(Image(img_bytes, width=400, height=180))
    wrapped_data = [data[0]]
    for row in data[1:]:
        wrapped_row = [Paragraph(str(cell), style_table) for cell in row]
        wrapped_data.append(wrapped_row)
    from reportlab.platypus import Table
    table = Table(wrapped_data, repeatRows=1, colWidths=[80, 200, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1976D2')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.whitesmoke, colors.lightblue]),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elements.append(table)
    doc = SimpleDocTemplate(io.BytesIO(), pagesize=landscape(letter), leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20)
    doc.build(elements)
    output = doc.filename
    output.seek(0)
    from flask import send_file
    return send_file(output, download_name='reporte_resumen.pdf', as_attachment=True)

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)